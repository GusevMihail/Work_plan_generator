from collections import namedtuple

from openpyxl.styles import NamedStyle, Border, Alignment, PatternFill, Font
from openpyxl.worksheet import Worksheet

TableArea = namedtuple('table_area', 'first_row first_col last_row last_col')


def apply_named_style(worksheet: Worksheet, style: NamedStyle, table_area: TableArea):
    for row in worksheet.iter_rows(min_row=table_area.first_row, max_row=table_area.last_row,
                                   min_col=table_area.first_col, max_col=table_area.last_col):
        for cell in row:
            cell.style = style


def apply_style(worksheet: Worksheet, table_area: TableArea, border: Border = None, alignment: Alignment = None,
                fill: PatternFill = None, font: Font = None):
    for row in worksheet.iter_rows(min_row=table_area.first_row, max_row=table_area.last_row,
                                   min_col=table_area.first_col, max_col=table_area.last_col):
        for cell in row:
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
