from abc import ABCMeta, abstractmethod
from collections import namedtuple
from typing import List

from openpyxl.worksheet import Worksheet

from Cell_styler import TableArea

import Pre_processing


def xstr(cell_value):
    if cell_value is None:
        return None
    else:
        return str(cell_value)


def xint(cell_value):
    if cell_value is None:
        return None
    else:
        return int(cell_value)


RawData = namedtuple('raw_data', 'day work_type place')


class AbstractParser(metaclass=ABCMeta):

    def __init__(self, sheet: Worksheet):
        self.sheet = sheet
        self.month_year = None
        self.raw_data: List[RawData] = []
        self.system = None

    @abstractmethod
    def _find_data_boundaries(self):
        raise NotImplementedError

    @abstractmethod
    def _find_month_year(self):
        raise NotImplementedError

    @abstractmethod
    def _extract_jobs(self):
        raise NotImplementedError


class ParserAsu(AbstractParser):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self._data_area = None
        self.system = Pre_processing.find_system_by_sheet(sheet.title)
        self._find_data_boundaries()
        self._find_month_year()
        self._extract_jobs()

    def _find_data_boundaries(self):
        max_table_row = 500
        max_table_col = 60
        first_row = None
        first_col = None
        last_row = None
        last_col = None

        for row in range(1, max_table_row):
            if xstr(self.sheet.cell(row, 1).value) == '1':
                for col in range(1, max_table_col):
                    if xstr(self.sheet.cell(row - 1, col).value) == '1':
                        first_row = row
                        first_col = col
                        break
                break

        for row in range(first_row, max_table_row):
            object_name_col = 2
            cell = self.sheet.cell(row, object_name_col).value
            if not isinstance(cell, str):
                last_row = row - 1
                break

        for col in range(first_col, max_table_col):
            if self.sheet.cell(first_row - 1, col).value is None:
                last_col = col - 1
                break
        self._data_area = TableArea(first_row, first_col, last_row, last_col)

    def _find_month_year(self):
        if self._data_area is None:
            self._find_data_boundaries()
        self.month_year = self.sheet.cell(self._data_area.first_row - 2, self._data_area.first_col).value
        if self.month_year is None:
            self.month_year = self.sheet.cell(self._data_area.first_row - 3, self._data_area.first_col).value

    def _extract_jobs(self):
        if self._data_area is None:
            self._find_data_boundaries()
        for i_row in range(self._data_area.first_row, self._data_area.last_row + 1):
            place_col = 2
            raw_place = self.sheet.cell(i_row, place_col).value
            for i_col in range(self._data_area.first_col, self._data_area.last_col + 1):
                raw_work_type = self.sheet.cell(i_row, i_col).value
                if raw_work_type is not None:
                    raw_day = self.sheet.cell(self._data_area.first_row - 1, i_col).value
                    i_raw_data = RawData(raw_day, raw_work_type, raw_place)
                    # print(i_raw_data)  # debug
                    self.raw_data.append(i_raw_data)


class ParserVOLS(AbstractParser):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = 'ВОЛС'
        self._place_in_header: str = None
        self._work_type_col = 4
        self._data_first_col = 7
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: int = None

        self._find_data_boundaries()
        self._find_month_year()
        self._find_place_in_header()
        self._extract_jobs()

    def _find_data_boundaries(self):
        max_table_row = 200
        max_table_col = 60

        for row in range(1, max_table_row):
            cell = str(self.sheet.cell(row, self._work_type_col).value)
            if ('ТО' in cell) and not ('вид' in cell.lower()):
                # print(f'[D{row}] {cell}')  # debug
                self._data_rows.append(row)
        self._data_rows.sort()

        for row in range(1, max_table_row):
            # print(row)  # debug
            cell = str(self.sheet.cell(row, self._data_first_col).value)
            if cell == '1':
                self._days_row = row
                break

        for col in range(self._data_first_col, max_table_col):
            cell = xint(self.sheet.cell(self._days_row, col).value)
            # print(f'[{col},{row}]={cell} ({type(cell)})')  # debug
            if type(cell) is not int:
                self._data_last_col = col - 1
                break

        # print(f'data boundaries: first col {self._data_first_col}, '
        #       f'last col {self._data_last_col}, rows {self._data_rows}')  # debug

    def _find_month_year(self):
        if self._days_row is None:
            self._find_data_boundaries()

        self.month_year = self.sheet.cell(self._days_row - 1, self._data_first_col).value
        # print(f'month_year = {self.month_year}')  # debug

    def _find_place_in_header(self):
        place_max_row = 30
        place_col = 1

        for row in range(1, place_max_row):
            cell = str(self.sheet.cell(row, place_col).value)
            if 'Местоположение' in cell:
                self._place_in_header = cell
                # print(f'place in header {self._place_in_header}')  # debug
                break

    def _find_place(self, data_row) -> str:
        if self._place_in_header is None:
            self._find_place_in_header()
        for row in range(data_row, self._days_row, -1):
            cell = str(self.sheet.cell(row, self._data_first_col).value)
            if 'ПС' in cell and '86' in cell:
                return 'ПС №86'
        else:
            return self._place_in_header

    def _extract_jobs(self):
        if self._data_last_col is None:
            self._find_data_boundaries()

        for row in self._data_rows:
            work_type = self.sheet.cell(row, self._work_type_col).value
            place = self._find_place(row)
            for col in range(self._data_first_col, self._data_last_col + 1):
                cell = self.sheet.cell(row, col).value
                if cell is not None:
                    day = xint(self.sheet.cell(self._days_row, col).value)
                    i_raw_data = RawData(day, work_type, place)
                    # print(i_raw_data)  # debug
                    self.raw_data.append(i_raw_data)
