from abc import ABCMeta, abstractmethod
from collections import namedtuple
from typing import List, Optional

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import pre_processing
from cell_styler import TableArea
from pre_processing import Systems


def xstr(cell_value):
    if cell_value is None:
        return None
    else:
        filtered_str = str(cell_value).strip(' \t\n')
        if filtered_str == '':
            return None
        else:
            return filtered_str


def xint(cell_value):
    if cell_value is None:
        return None
    elif type(cell_value) is str:
        return pre_processing.find_num_in_str(cell_value)
    else:
        return cell_value


RawData = namedtuple('raw_data', 'day work_type place tech_map equip_name')


class AbstractParser(metaclass=ABCMeta):

    def __init__(self, sheet: Worksheet):
        self.sheet = sheet
        self.month_year = None
        self.raw_data: List[RawData] = []
        self.system = None

    def get_cell(self, row, col):
        return self.sheet.cell(row, col).value

    @abstractmethod
    def _find_data_boundaries(self):
        raise NotImplementedError

    @abstractmethod
    def _find_month_year(self):
        raise NotImplementedError

    @abstractmethod
    def _extract_jobs(self):
        raise NotImplementedError


class ParserAsu(AbstractParser):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self._data_area = None
        self.system = pre_processing.find_system_by_sheet(sheet.title)
        self._find_data_boundaries()
        self._find_month_year()
        self._extract_jobs()

    def _find_data_boundaries(self):
        max_table_row = 500
        max_table_col = 60
        first_row = None
        first_col = None
        last_row = None
        last_col = None

        for row in range(1, max_table_row):
            if xstr(self.sheet.cell(row, 1).value) == '1':
                for col in range(1, max_table_col):
                    if xstr(self.sheet.cell(row - 1, col).value) == '1':
                        first_row = row
                        first_col = col
                        break
                break

        for row in range(first_row, max_table_row):
            object_name_col = 2
            cell = self.sheet.cell(row, object_name_col).value
            if not isinstance(cell, str):
                last_row = row - 1
                break

        for col in range(first_col, max_table_col):
            if self.sheet.cell(first_row - 1, col).value is None:
                last_col = col - 1
                break
        self._data_area = TableArea(first_row, first_col, last_row, last_col)

    def _find_month_year(self):
        if self._data_area is None:
            self._find_data_boundaries()
        self.month_year = self.sheet.cell(self._data_area.first_row - 2, self._data_area.first_col).value
        if self.month_year is None:
            self.month_year = self.sheet.cell(self._data_area.first_row - 3, self._data_area.first_col).value

    def _extract_jobs(self):
        if self._data_area is None:
            self._find_data_boundaries()
        for i_row in range(self._data_area.first_row, self._data_area.last_row + 1):
            place_col = 2
            raw_place = self.sheet.cell(i_row, place_col).value
            for i_col in range(self._data_area.first_col, self._data_area.last_col + 1):
                raw_work_type = self.sheet.cell(i_row, i_col).value
                if raw_work_type is not None:
                    raw_day = self.sheet.cell(self._data_area.first_row - 1, i_col).value
                    one_line_work_type = xstr(raw_work_type).replace('\n', ' ')
                    for splitted_work_type in one_line_work_type.split(' '):
                        i_raw_data = RawData(raw_day, splitted_work_type, raw_place, '',
                                             '')  # старый парсер, работа с техкартой и оборудованием не поддерживается
                        if i_raw_data not in self.raw_data:
                            self.raw_data.append(i_raw_data)


class ParserVolsLikeSys(AbstractParser):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = None
        self._place_in_header: Optional[str] = None
        self._work_type_col = 4
        self._data_first_col = 7
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: Optional[int] = None

    def _find_data_boundaries(self):
        max_table_row = 200
        max_table_col = 60

        # find data rows
        for row in range(1, max_table_row):
            cell = str(self.sheet.cell(row, self._work_type_col).value)
            row_visible = not self.sheet.row_dimensions[row].hidden
            if ('ТО' in cell) and not ('вид' in cell.lower()) and row_visible:
                self._data_rows.append(row)
        self._data_rows.sort()

        # find days row and first data col
        exit_flag = False
        for row in range(1, max_table_row):
            row_visible = not self.sheet.row_dimensions[row].hidden
            if row_visible:
                for col in range(self._work_type_col, 60):
                    column_visible = not self.sheet.column_dimensions[get_column_letter(col)].hidden
                    cell = xstr(self.sheet.cell(row, col).value)
                    next_cell = xstr(self.sheet.cell(row, col + 1).value)
                    if column_visible and cell == '1' and next_cell == '2':
                        self._days_row = row
                        self._data_first_col = col
                        exit_flag = True
                        break
            if exit_flag:
                break

        for col in range(self._data_first_col, max_table_col):
            cell_value = xint(self.sheet.cell(self._days_row, col).value)
            column_hidden = self.sheet.column_dimensions[get_column_letter(col)].hidden
            # print(f'[{col},{self._days_row}]={cell} ({type(cell)})')  # debug
            if type(cell_value) is not int or column_hidden:
                self._data_last_col = col - 1
                break

        # print(f'data boundaries: first col {self._data_first_col}, '
        #       f'last col {self._data_last_col}, rows {self._data_rows}')  # debug

    def _find_month_year(self):
        if self._days_row is None:
            self._find_data_boundaries()

        self.month_year = self.sheet.cell(self._days_row - 1, self._data_first_col).value
        # print(f'month_year = {self.month_year}')  # debug

    def _find_place_in_header(self):
        place_max_row = 30
        place_col = 1

        for row in range(1, place_max_row):
            cell = str(self.sheet.cell(row, place_col).value)
            if 'Место' in cell:
                self._place_in_header = cell
                # print(f'place in header {self._place_in_header}')  # debug
                break

    def _find_place(self, data_row) -> str:
        if self._place_in_header is None:
            self._find_place_in_header()

        for row in range(data_row, self._days_row, -1):
            place_col = 1
            # place_col = self._data_first_col
            cell = str(self.sheet.cell(row, place_col).value)
            place, object_name = pre_processing.extract_place_and_object(cell)
            if object_name != 'unknown':
                # print(f'{cell.ljust(40)} -> place {place}, obj {object_name}')
                return place
        else:
            return self._place_in_header

    def _extract_jobs(self):
        for row in self._data_rows:
            work_type = self.sheet.cell(row, self._work_type_col).value
            place = self._find_place(row)
            for col in range(self._data_first_col, self._data_last_col + 1):
                cell = xstr(self.sheet.cell(row, col).value)
                if cell is not None:
                    day = xint(self.sheet.cell(self._days_row, col).value)
                    i_raw_data = RawData(day, work_type, place)
                    if i_raw_data not in self.raw_data:
                        self.raw_data.append(i_raw_data)


class ParserSake(AbstractParser):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)

        self._place_in_data_area_col = 2
        self._place_in_header_row = 4
        self._document_id_row = 2

        self._equip_name_col = 3
        self._tech_map_col = 5
        self._work_type_col = 9
        self._data_first_col = 12
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: Optional[int] = 19
        self._date_col = 2
        self._date_row = None
        self._sys_row = None

        self.month_year = self._find_month_year()

        cell_system = self.get_cell(self._sys_row, self._date_col)  # system and date has same column number
        self.system = self.str_to_system(cell_system)

        # self.document_id = self.get_cell(2, 40)
        self._find_place_in_header()
        self._get_is_multiplace()

        self._find_data_boundaries()
        self._extract_jobs()

    def _find_month_year(self):
        for row in range(10, 25):
            cell = self.get_cell(row, self._date_col)
            if cell:
                self._date_row = row
                self._sys_row = row + 2
                return cell

    @staticmethod
    def str_to_system(sys_str: str):
        aliases = {'АИИСКУЭ': Systems.ASKUE,
                   'учет': Systems.TECH_REG,
                   'М2': Systems.TK,
                   'ВОЛС': Systems.VOLS,
                   'ЛВС': Systems.LVS,
                   'АСУ ТП': Systems.ASU_TP,
                   'АСУ И': Systems.ASU_I,
                   'мост': Systems.ASU_AM}

        for alias, system in aliases.items():
            if alias in sys_str:
                return system
        else:
            return None

    def _find_data_boundaries(self):
        first_table_row = 22
        max_table_row = 120
        max_table_col = 50

        # find data rows
        for row in range(first_table_row, max_table_row):
            cell = str(self.get_cell(row, self._work_type_col))
            row_visible = not self.sheet.row_dimensions[row].hidden
            if ('ТО' in cell) and not ('вид' in cell.lower()) and row_visible:
                self._data_rows.append(row)
        self._data_rows.sort()

        # find data last column
        for col in range(self._data_first_col, max_table_col):
            cell_value = xint(self.get_cell(self._days_row, col))
            column_hidden = self.sheet.column_dimensions[get_column_letter(col)].hidden
            # print(f'[{col},{self._days_row}]={cell} ({type(cell)})')  # debug
            if type(cell_value) is not int or column_hidden:
                self._data_last_col = col - 1
                break

        # print(f'data boundaries: first col {self._data_first_col}, '
        #       f'last col {self._data_last_col}, rows {self._data_rows}')  # debug

    def _get_is_multiplace(self):
        multiplace_documents = ('10.4.36',  # АИИСКУЭ для С1\Бронка ПС86\Котлин
                                '10.4.38',  # ВОЛС для С1\Бронка ПС86\Котлин
                                '13.0.107',  # АСУ ТП
                                '14.0.107',  # АСУ ТП
                                '13.0.108',  # АСУ И
                                '14.0.108',  # АСУ И
                                '2.1.109',  # АСУ АМ
                                '13.0.110',  # ЛВС
                                '14.0.110')  # ЛВС
        self.is_multiplace = self.document_id in multiplace_documents

    def _extract_jobs(self):
        for row in self._data_rows:
            work_type = self.get_cell(row, self._work_type_col)
            tech_map = self.get_cell(row, self._tech_map_col)
            equip_name = self.get_cell(row, self._equip_name_col)

            if self.is_multiplace:
                place = self._find_place_in_data_area(row)

            else:
                place = self.place_in_header

            for col in range(self._data_first_col, self._data_last_col + 1):
                cell = xstr(self.get_cell(row, col))
                if cell is not None:
                    day = xint(self.get_cell(self._days_row, col))

                    # заменяем place С1ПС110/10 -> ПС86 для ВОЛС по номеру техкарты
                    ps86_vols_tech_maps = (
                        'Технологическая карта 3/2/2016',
                        'Технологическая карта 15/2/2016',
                        'Технологическая карта 16/2/2016',
                        'Технологическая операция 3/2/2016',
                        'Технологическая операция 15/2/2016',
                        'Технологическая операция 16/2/2016',
                    )
                    if self.system == Systems.VOLS and (tech_map in ps86_vols_tech_maps):
                        place = pre_processing.Objects.PS86.value

                    i_raw_data = RawData(day, work_type, place, tech_map, equip_name)
                    if i_raw_data not in self.raw_data:
                        self.raw_data.append(i_raw_data)

    def _find_place_in_header(self):
        columns_to_search = range(30, 55)
        template = 'Объект имущества'
        for col in columns_to_search:
            cell = self.get_cell(self._place_in_header_row, col)
            if cell and (template in cell):
                self.place_in_header = self.get_cell(self._place_in_header_row, col + 4)
                self.document_id = self.get_cell(self._document_id_row, col + 4)
                break
        else:
            raise Exception(f'Cant find place in header')

    def _find_place_in_data_area(self, data_row):
        for row in range(data_row, self._days_row, -1):
            cell = str(self.get_cell(row, self._place_in_data_area_col))
            place, object_name = pre_processing.extract_place_and_object(cell)

            if self.document_id == '10.4.38' and 'С1' in place:
                place = 'C1 ПС 110/10кВ'
            if self.document_id in ('14.0.107',  # АСУ ТП
                                    '14.0.108',  # АСУ И
                                    '2.1.109',  # АСУ АМ
                                    '14.0.110'  # ЛВС
                                    ) \
                    and 'Котлин' in place:
                place = pre_processing.Objects.ZU.value
                object_name = pre_processing.Objects.ZU

            if object_name != 'unknown':
                return place
        else:
            raise Exception(f'Can\'t find place in data area for row {data_row},'
                            f'sheet {self.sheet.title},'
                            f'system {self.system}')


class ParserVols(ParserVolsLikeSys):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = Systems.VOLS  # unique value
        self._place_in_header: Optional[str] = None
        self._work_type_col = 4
        self._data_first_col = 7  # unique value
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: Optional[int] = None

        self._find_data_boundaries()
        self._find_month_year()
        self._find_place_in_header()
        self._extract_jobs()


class ParserVols_v2(ParserSake):

    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = Systems.VOLS
        self._place_in_header: Optional[str] = None
        self._place_in_data_area_col = 2
        self._work_type_col = 9
        self._data_first_col = 12
        self._days_row: Optional[int] = 19
        self.date_cell = (12, 2)
        self._data_rows: List[int] = []

        self._find_data_boundaries()
        self._find_month_year()
        self._extract_jobs()

    def _find_month_year(self):
        self.month_year = self.sheet.cell(*self.date_cell).value
        # print(f'month_year = {self.month_year}')  # debug

    def _find_place_in_data_area(self, data_row):
        for row in range(data_row, self._days_row, -1):
            cell = str(self.sheet.cell(row, self._place_in_data_area_col).value)
            place, object_name = pre_processing.extract_place_and_object(cell)
            if object_name != 'unknown':
                return place
        else:
            raise Exception(f'Cant find place in data area for row {data_row},'
                            f'sheet {self.sheet.title},'
                            f'system {self.system}')

    def _find_place(self, data_row) -> str:
        return self._find_place_in_data_area(data_row)


class ParserTk(ParserVolsLikeSys):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = Systems.TK  # unique value
        self._place_in_header: Optional[str] = None
        self._work_type_col = 4
        self._data_first_col = 7  # unique value
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: Optional[int] = None

        self._find_data_boundaries()
        self._find_month_year()
        self._find_place_in_header()
        self._extract_jobs()

    def _find_month_year(self):
        date_cell = (3, 1)

        self.month_year = self.sheet.cell(*date_cell).value
        # print(f'month_year = {self.month_year}')  # debug


class ParserAskue(ParserVolsLikeSys):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = Systems.ASKUE  # unique value
        self._place_in_header: Optional[str] = None
        self._work_type_col = 4
        self._data_first_col = 6  # unique value
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: Optional[int] = None

        self._find_data_boundaries()
        self._find_month_year()
        self._find_place_in_header()
        self._extract_jobs()


class ParserTechReg(ParserVolsLikeSys):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = Systems.TECH_REG  # unique value
        self._place_in_header: Optional[str] = None
        self._work_type_col = 6  # unique value
        self._data_first_col = 9  # unique value
        self._data_last_col = None
        self._data_rows: List[int] = []
        self._days_row: Optional[int] = None

        self._find_data_boundaries()
        self._find_month_year()
        self._find_place_in_header()
        self._extract_jobs()


class ParserAskueSake(ParserSake):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)
        self.system = Systems.ASKUE
        self._place_in_header_row = 4
        self._place_in_header_col = 39
        self._place_in_data_area_col = 2
        self._work_type_col = 9
        self._data_first_col = 12
        self._days_row = 19
        self._year_row = 12
        self._year_col = 2

        self._find_data_boundaries()
        self._find_month_year()
        self._find_place_in_header()
        self._extract_jobs()
