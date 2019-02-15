from typing import List
from itertools import groupby

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

from Cell_styler import apply_style, TableArea
from Pre_processing import Job


class WorkPlan:

    def __init__(self, jobs: List[Job], template_path: str):
        self.jobs = jobs
        self._first_col = None
        self._last_col = None
        self._first_data_row = None
        self._current_row = None
        self._thin_side = Side(style='thin', color='000000')
        self._thin_border = Border(left=self._thin_side, top=self._thin_side,
                                   right=self._thin_side, bottom=self._thin_side)
        self._basic_font = Font(name='Times New Roman', size=10)
        self._bold_font = Font(name='Times New Roman', size=11, b=True)
        self._align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self._wb = self._get_template(template_path)
        self._ws = self._wb.active
        self._write_date()

    def _get_template(self, filename: str):
        wb_template = openpyxl.load_workbook(filename)
        ws_template = wb_template.active
        self._first_col = 1
        self._last_col = 9
        self._first_data_row = 10
        self._current_row = self._first_data_row
        table_header = TableArea(first_row=8, last_row=9, first_col=self._first_col, last_col=self._last_col)
        apply_style(ws_template, table_header, border=self._thin_border)
        return wb_template

    def _write_date(self):
        date_cell = self._ws.cell(row=6, column=1)
        date_cell.value = self.jobs[0].date

    def _write_obj_row(self, obj_name):
        self._ws.merge_cells(start_row=self._current_row, end_row=self._current_row,
                             start_column=self._first_col, end_column=self._last_col)

        current_row_area = TableArea(self._current_row, self._first_col, self._current_row, self._last_col)
        apply_style(self._ws, current_row_area,
                    border=self._thin_border,
                    font=self._bold_font,
                    alignment=self._align_center)

        self._ws.cell(self._current_row, self._first_col).value = obj_name
        self._current_row += 1

    def _write_work_row(self, job: Job):
        current_row_area = TableArea(self._current_row, self._first_col, self._current_row, self._last_col)
        apply_style(self._ws, current_row_area,
                    border=self._thin_border,
                    font=self._basic_font,
                    alignment=self._align_center)
        organization_col = 1
        organization = 'ООО "Би.Си.Си."'
        system_col = 2
        work_col = 3
        place_col = 4
        work_start_col = 5
        work_start = '9:00'
        work_end_col = 6
        work_end = '18:00'
        worker_col = 7
        self._ws.cell(self._current_row, organization_col).value = organization
        self._ws.cell(self._current_row, system_col).value = job.system
        self._ws.cell(self._current_row, work_col).value = job.work_type  # + ' ' + job.system
        self._ws.cell(self._current_row, place_col).value = job.place
        self._ws.cell(self._current_row, work_start_col).value = work_start
        self._ws.cell(self._current_row, work_end_col).value = work_end
        self._ws.cell(self._current_row, worker_col).value = job.worker
        self._current_row += 1

    def make_plan(self):
        jobs_by_object = groupby(self.jobs, key=lambda i_job: i_job.object)
        for object_name, object_jobs in jobs_by_object:
            self._write_obj_row(object_name)
            for job in object_jobs:
                self._write_work_row(job)

    def save_file(self):
        self._wb.save(f'.\\output data\\{self.jobs[0].date}.xlsx')
        self._wb.close()

# if __name__ == '__main__':
