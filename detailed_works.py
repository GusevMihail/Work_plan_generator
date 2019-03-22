import openpyxl
from openpyxl.workbook.workbook import Workbook, Worksheet
from Pre_processing import Job


def work_details(job: Job) -> str:
    workbook = openpyxl.load_workbook(r'.\input data\Detailed works.xlsx')
    sheet = workbook.active
    max_row = 1000
    system_col = 1
    work_type_col = 2
    details_col = 3

    for row in range(1, max_row + 1):
        system = sheet.cell(row, system_col).value
        work_type = sheet.cell(row, work_type_col).value
        details = sheet.cell(row, details_col).value
        if job.system == system and job.work_type == work_type:
            return details
    return None
