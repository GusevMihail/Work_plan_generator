from openpyxl.utils import get_column_letter
from openpyxl.worksheet import Worksheet
from Parser import xint
from collections import namedtuple

worker = namedtuple('worker', 'full_name phone_number')
class DutySchedule:
    def __init__(self, worksheet: Worksheet):
        self.worksheet = worksheet
        self._data_first_col = 3
        self._data_last_col = None
        self._find_last_day_col()
        # номера строк вносятся в порядке убывания приоритета сотрудника для выбора его в качестве руководителя работ
        self.group_s1_rows = (8, 9, 10, 11)
        self.group_s2_rows = (14, 13)
        self.group_v_rows = (17, 18, 19)
        self.group_vols_rows = (23, 21, 22)
        self.group_tk_rows = (22, 21, 23)
        self.all_workers_rows = self.group_s1_rows + self.group_s2_rows + self.group_v_rows + \
                                self.group_vols_rows + self.group_tk_rows

        self.workers =

    def _find_last_day_col(self):
        days_row = 2
        table_max_col = 50
        for col in range(self._data_first_col, table_max_col + 1):
            cell = self.worksheet.cell(days_row, col).value
            if type(xint(cell)) is not int:
                self._data_last_col = col - 1
                break
        else:
            print('Последняя дата не найдена')

    def _day2col(self, day: int):
        col = day + self._data_first_col - 1
        if col <= self._data_last_col:
            return col
        else:
            return None

    def _is_workday(self, day: int):
        col = self._day2col(day)
        if col is None:
            return None
        else:
            for row in self.all_workers_rows:
                cell = xint(self.worksheet.cell(row, col).value)
                if cell == 8 or cell == 7:
                    return True
            else:
                return False

    # def get_s1_worker(self, day: int):
    #     for row in self.group_s1_rows
    #         if
