from abc import ABCMeta, abstractmethod
from typing import List, Tuple, Union, Type

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
import pandas as pd

from pre_processing import Job, Systems, Objects


def jobs_to_df(jobs: List[Job]) -> pd.DataFrame:
    columns = ('date', 'system', 'object', 'place', 'work_type', 'tech_map', 'equip_name', 'performer',)
    result = pd.DataFrame(columns=columns,
                          data=((j.date, j.system, j.object, j.place, j.work_type,
                                 j.tech_map, j.equip_name, j.performer)
                                for j in jobs))
    return result


class Journal(metaclass=ABCMeta):
    def __init__(self, jobs_df: pd.DataFrame):
        self.df = jobs_df
        self.journal: Union[None, pd.DataFrame] = None
        self.default_header = None

    @abstractmethod
    def make_journal(self, sys: Systems, obj: Objects, place: str):
        raise NotImplementedError

    def save_journal(self, file_name='default', folder=r'./output data/journals/', header='default'):
        if self.journal is None:
            raise ValueError('make journal before save it!')
        else:
            full_file_name = folder + self.df.date[0].strftime('%Y %m ') + file_name + '.xlsx'
            sheet_name = 'Sheet1'
            if header == 'default':
                header = self.default_header
            header_names = tuple(col[0] for col in header)
            col_widths = tuple(col[1] for col in header)
            with pd.ExcelWriter(path=full_file_name, date_format='DD.MM.YY') as writer:
                self.journal.to_excel(writer, sheet_name, header=header_names, index=False)
                self._set_wrap_text(writer.sheets[sheet_name], (2,))
                self._set_columns_width(writer.sheets[sheet_name], col_widths)
                self._set_print_area(writer.sheets[sheet_name])
                self._clear_dates(writer.sheets[sheet_name])

    def _set_wrap_text(self, ws: Worksheet, col_numbers: Tuple[int]):
        table_last_row = self.journal.shape[0] + 1
        for row in range(1, table_last_row + 1):
            for col in col_numbers:
                ws.cell(row, col).alignment = Alignment(wrap_text=False)
        for row in range(2, table_last_row + 1):
            date_col = 1
            ws.cell(row, date_col).alignment = Alignment(horizontal='left')

    @staticmethod
    def _set_columns_width(ws: Worksheet, col_widths):
        from openpyxl.utils import get_column_letter
        for i, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width

    # @staticmethod
    def _set_print_area(self, ws: Worksheet):
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.scale = 75

    def _clear_dates(self, ws: Worksheet):
        table_last_row = self.journal.shape[0] + 1
        date_col = 1
        prev_date = ws.cell(row=2, column=date_col).value
        for row in range(3, table_last_row + 1):
            date = ws.cell(row, date_col).value
            if date == prev_date:
                ws.cell(row, date_col).value = ''
            prev_date = date


class JournalASU(Journal):
    def __init__(self, jobs_df: pd.DataFrame):
        super().__init__(jobs_df)
        from config_journals import default_header_ASU
        self.default_header = default_header_ASU

    def make_journal(self, sys: Systems, obj: Objects = None, place_filter: str = None):
        if obj:
            journal = self.df[self.df.object == obj]
        else:
            journal = self.df
        journal = journal[journal.system == sys] \
            .drop(['system', 'object'], axis=1) \
            .reindex(columns=['date', 'place', 'work_type', 'tech_map', 'performer']) \
            .sort_values(by=['date', 'place', 'work_type', 'tech_map']) \
            .drop_duplicates()
        if place_filter:
            journal = journal[journal.place.str.contains(place_filter)]
        journal.performer = journal.performer.apply(lambda w: w.last_name)
        self.journal = journal


class JournalASKUE(Journal):
    def __init__(self, jobs_df: pd.DataFrame):
        super().__init__(jobs_df)
        from config_journals import default_header_ASKUE
        self.default_header = default_header_ASKUE

    def make_journal(self, sys: Systems, obj: Objects = None, place_filter: str = None):
        if obj:
            journal = self.df[self.df.object == obj]
        else:
            journal = self.df
        journal = journal[journal.system == sys] \
            .drop(['object'], axis=1) \
            .reindex(columns=['date', 'place', 'equip_name', 'work_type', 'tech_map', 'performer']) \
            .sort_values(by=['date', 'place', 'work_type', 'tech_map'])
        if place_filter:
            journal = journal[journal.place.str.contains(place_filter)]
        journal.performer = journal.performer.apply(lambda w: w.last_name)
        self.journal = journal


def batch_journal_generator(jobs_df: pd.DataFrame,
                            journal_class: Type[Journal],
                            config: dict,
                            return_journals=True,
                            save_journals=True,
                            verbose=False):
    journals = {}
    for name, conf in config.items():
        if verbose:
            print(name)
        sys, obj, place_filter = conf
        j = journal_class(jobs_df)
        j.make_journal(sys, obj, place_filter)
        if return_journals:
            journals[name] = j
        if save_journals:
            j.save_journal(name)
    return journals
