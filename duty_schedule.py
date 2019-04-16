from datetime import date
from enum import Enum
from typing import List, Iterable, Union

from openpyxl.worksheet.worksheet import Worksheet

from pre_processing import extract_month_and_year, Systems, Objects


class Worker:
    def __init__(self, last_name: str, first_name: str, patronymic: str, phone: str):
        self.last_name = last_name
        self.first_name = first_name
        self.patronymic = patronymic
        self.phone = phone

    def __str__(self):
        return f'{self.last_name} {self.first_name} {self.patronymic} {self.phone} '

    def __repr__(self):
        return self.__str__()


class Team:
    def __init__(self, name: Union[None, str, Systems, Objects] = None, workers: List[Worker] = None):
        self.name = name
        if workers is None:
            self.workers = []
        else:
            self.workers = workers

    def __str__(self):
        return self.name + ': ' + ', '.join(w.last_name for w in self.workers)

    def __repr__(self):
        return self.__str__()

    def get_by_last_name(self, last_name: str) -> Worker:
        for w in self.workers:
            if w.last_name in last_name:
                return w

    def get_by_last_names(self, last_names: Iterable[str]) -> List[Worker]:
        return [self.get_by_last_name(name) for name in last_names]


# workers directory
all_workers = Team('Все сотрудники', [
    Worker('Борисевич', 'Константин', 'Васильевич', '+79215943419'),
    Worker('Добрицкий', 'Дмитрий', 'Александрович', '+79315347202'),
    Worker('Харитонов', 'Виктор', 'Яковлевич', '+79319642368'),
    Worker('Гусев', 'Михаил', 'Владимирович', '+79675904368'),
    Worker('Мулин', 'Николай', 'Николаевич', '+79112283889'),
    Worker('Кушмылев', 'Евгений', 'Павлович', '+79819134737'),
    Worker('Каприца', 'Анатолий', 'Евгеньевич', '+79046315018'),
    Worker('Макаров', 'Виктор', 'Викторович', '+79312089129'),
    Worker('Горнов', 'Александр', 'Серафимович', '+79111314015'),
    Worker('Подольский', 'Андрей', 'Вениаминович', '+79312531066'),
    Worker('Кокоев', 'Михаил', 'Николаевич', '+79216441993'),
    Worker('Санжара', 'Владимир', 'Александрович', '+79819629285'),
    Worker('Ильин', 'Андрей', 'Владимирович', '+79219303652'),
    Worker('Ястребов', 'Алексей', 'Владимирович', '+79313581975'),
    Worker('Огородников', 'Алексей', 'Юрьевич', '+79313196196')
])

team_heads = Team('Руководители', all_workers.get_by_last_names(('Борисевич', 'Добрицкий')))
team_s1 = Team(Objects.S1, all_workers.get_by_last_names(('Харитонов', 'Гусев', 'Мулин', 'Кушмылев')))
team_s2 = Team(Objects.S2, all_workers.get_by_last_names(('Каприца', 'Горнов')))
team_v = Team('В1-В6', all_workers.get_by_last_names(('Подольский', 'Кокоев', 'Санжара')))
team_vols = Team(Systems.VOLS, all_workers.get_by_last_names(('Ильин', 'Ястребов', 'Огородников')))
team_tk = Team(Systems.TK, all_workers.get_by_last_names(('Огородников', 'Ястребов', 'Ильин')))
team_askue = Team(Systems.ASKUE, all_workers.get_by_last_names(('Ястребов', 'Огородников', 'Ильин')))


class WorkerStatus(Enum):
    ON_WORK = 'На работе'
    NOT_ON_WORK = 'Не на работе'
    ON_DUTY = 'На дежурстве'


class DutySchedule:
    def __init__(self, worksheet: Worksheet, _all_workers: Team):
        self.worksheet = worksheet
        self.all_workers = _all_workers
        self.month, self.year = extract_month_and_year(worksheet.cell(1, 3).value)
        self.workers_rows = range(1, 50)
        self.workers_col = 2

    @staticmethod
    def _date2col(target_date: date) -> int:
        date_first_col = 3
        return date_first_col - 1 + target_date.day

    def _find_worker_row(self, worker: Worker) -> int:
        for row in self.workers_rows:
            cell = self.worksheet.cell(row, self.workers_col).value
            if worker.last_name in str(cell).replace('ё', 'е'):
                return row

    def get_worker_status(self, worker: Worker, target_date: date) -> WorkerStatus:
        cell = self.worksheet.cell(self._find_worker_row(worker), self._date2col(target_date)).value
        if cell in (8, 7):
            return WorkerStatus.ON_WORK
        elif cell == 15:
            return WorkerStatus.ON_DUTY
        else:
            return WorkerStatus.NOT_ON_WORK

    def is_workday(self, target_date: date) -> bool:
        col = self._date2col(target_date)
        first_worker_row = 5
        cell = self.worksheet.cell(first_worker_row, col).value
        return cell in (7, 8)

    def get_duty_str(self, target_date: date) -> str:
        for row in self.workers_rows:
            cell = self.worksheet.cell(row, self._date2col(target_date)).value
            if cell == 15:
                return str(self.worksheet.cell(row, self.workers_col).value).replace('ё', 'е')

    def get_performer(self, team: Team, target_date: date) -> Worker:
        for worker in team.workers:
            if self.get_worker_status(worker, target_date) == WorkerStatus.ON_WORK:
                return worker
        else:
            return self.all_workers.get_by_last_name(self.get_duty_str(target_date))

    def get_head(self, team: Team, target_date: date) -> Worker:
        for worker in team.workers:
            cell = self.worksheet.cell(self._find_worker_row(worker), self._date2col(target_date)).value
            if (type(cell) is int) or (cell is None) or (cell == ''):
                return worker
        else:
            return team.workers[0]


if __name__ == '__main__':
    print('')
    # print(all_workers)

    # print(team_s1)
    # print(team_s2)
    # print(team_v)
    # print(team_vols)
    # print(team_tk)
    # print(team_askue)
