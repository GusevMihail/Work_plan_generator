import pandas as pd
from openpyxl.utils import get_column_letter
import config_work_calendar


def concat_agg(iterable_value):
    #  функция для конкатенации
    #  используется при групировке полей в сводной таблице
    return ',\n'.join(iterable_value)


def make_calendar(pivot_table: pd.DataFrame,
                  description: config_work_calendar.CalendarDescription,
                  path=r'output data/calendars/',
                  sheet_name='Sheet1'):
    from datetime import datetime
    date = pivot_table.columns[0]
    date = datetime.strftime(date, '%Y %m')
    with pd.ExcelWriter(path=path + date + ' Календарь работ ' + description.name + '.xlsx',
                        date_format='DD.MM', datetime_format='DD.MM') as writer:
        pivot_table.to_excel(writer, sheet_name)
        ws = writer.sheets[sheet_name]
        ws.cell(row=1, column=1).value = description.name
        ws.column_dimensions[get_column_letter(1)].width = 35  # places column
        ws.column_dimensions[get_column_letter(2)].width = 17  # systems column
        for col in range(3, 35):
            ws.column_dimensions[get_column_letter(col)].width = 6
        from cell_styler import apply_style, TableArea
        from openpyxl.styles import Alignment
        apply_style(ws, TableArea(0, 0, 40, 40), alignment=Alignment(wrap_text=True))
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE


def calendar_pivot_table(df: pd.DataFrame, description: config_work_calendar.CalendarDescription) -> pd.DataFrame:
    if description.objects is not None:
        df = df[df.object.isin(description.objects)]
    if description.systems is not None:
        df = df[df.system.isin(description.systems)]
    df.system = df.system.apply(lambda i: i.value)
    df = df[['date', 'place', 'work_type', 'system']]
    df.drop_duplicates(inplace=True)
    pivot = df.pivot_table(values='work_type', index=['place', 'system'], columns='date', aggfunc=concat_agg)

    return pivot


def batch_make_calendars(df: pd.DataFrame):
    for description in config_work_calendar.calendars_settings:
        print(description.name)
        pivot = calendar_pivot_table(df, description)
        make_calendar(pivot, description)
