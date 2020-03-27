import datetime
import unittest
from typing import List

import openpyxl
from openpyxl.styles import NamedStyle

import application
import cell_styler
import duty_schedule
import pre_processing
import works_parser
from pre_processing import Systems, Objects


class TestParser(unittest.TestCase):

    def test_xstr(self):
        func = works_parser.xstr
        self.assertEqual(func('test str'), 'test str')
        self.assertEqual(func('   test str \n'), 'test str')
        self.assertEqual(func('123'), '123')
        self.assertEqual(func(123), '123')
        self.assertEqual(func(None), None)
        self.assertEqual(func(''), None)
        self.assertEqual(func('   '), None)
        self.assertEqual(func('\t'), None)
        self.assertEqual(func('\n   \t'), None)

    def test_xint(self):
        func = works_parser.xint
        self.assertEqual(func('1'), 1)
        self.assertEqual(func('  85  \n'), 85)
        self.assertEqual(func('str str 43 str'), 43)
        self.assertEqual(func('72, 234'), 72)
        self.assertEqual(func('float 1,32'), 1)
        self.assertEqual(func(''), None)
        self.assertEqual(func('   '), None)
        self.assertEqual(func('str str'), None)
        self.assertEqual(func(None), None)
        self.assertEqual(func([1, 2]), [1, 2])


class TestParserAsu(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb_asu = openpyxl.load_workbook(r'.\input data\5. Графики на 05.18 АСУ.xlsx')
        cls.wb_test = openpyxl.load_workbook(r'.\input data\Test Schedule.xlsx')
        cls.sheet_asu_109 = cls.wb_asu['109. МОСТ']
        cls.sheet_asu_107 = cls.wb_asu['107. АСУ ТП']
        cls.sheet_test_107 = cls.wb_test['107. АСУ ТП']
        cls.parser_asu_109 = works_parser.ParserAsu(cls.sheet_asu_109)
        cls.parser_asu_107 = works_parser.ParserAsu(cls.sheet_asu_107)
        cls.parser_test_107 = works_parser.ParserAsu(cls.sheet_test_107)

    def test_find_data_boundaries(self):
        self.assertEqual(self.parser_asu_109._data_area, (8, 6, 8, 36))
        self.assertEqual(self.parser_asu_107._data_area, (5, 3, 15, 33))

    def test_find_month_year(self):
        self.assertEqual(self.parser_asu_107.month_year, 'МАЙ  2018 г.')
        self.assertEqual(self.parser_asu_109.month_year, 'май 2018г.')

    def test_extract_jobs(self):
        last = len(self.parser_test_107.raw_data) - 1
        self.assertEqual(self.parser_test_107.raw_data[0].day, 1)
        self.assertEqual(self.parser_test_107.raw_data[0].work_type, 'ТО1')
        self.assertEqual(self.parser_test_107.raw_data[0].place, 'ПТК судопропускного сооружения - ПТК С1 север')
        self.assertEqual(self.parser_test_107.raw_data[last].day, 31)
        self.assertEqual(self.parser_test_107.raw_data[last].work_type, 'ТО4')
        self.assertEqual(self.parser_test_107.raw_data[last].place, 'ПТК ЗУ КЗС')


class TestParserVols(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb_vols = openpyxl.load_workbook(r'.\input data\Test Schedule VOLS.xlsx')
        cls.sheet_1 = cls.wb_vols['8.1.38 ТО']
        cls.sheet_2 = cls.wb_vols['10.4.38 ТО']
        cls.parser_1 = works_parser.ParserVols(cls.sheet_1)
        cls.parser_2 = works_parser.ParserVols(cls.sheet_2)

    def test_find_data_boundaries(self):
        self.assertEqual(self.parser_1._data_first_col, 7)
        self.assertEqual(self.parser_1._data_last_col, 37)
        self.assertEqual(self.parser_1._data_rows, [20, 21])
        self.assertEqual(self.parser_2._data_first_col, 7)
        self.assertEqual(self.parser_2._data_last_col, 37)
        self.assertEqual(self.parser_2._data_rows, [22, 23, 26])

    def test_find_month_year(self):
        self.assertEqual(self.parser_1.month_year, 'Май 2018 года')
        self.assertEqual(self.parser_2.month_year, 'Май 2018 года')

    def test_extract_jobs(self):
        self.assertEqual(self.parser_1.raw_data[0].day, 11)
        self.assertEqual(self.parser_1.raw_data[0].work_type, 'ТО2')
        self.assertEqual(self.parser_1.raw_data[0].place,
                         'Местоположение: Здание управления комплекса защитных сооружений')
        self.assertEqual(self.parser_1.raw_data[1].day, 18)
        self.assertEqual(self.parser_1.raw_data[1].work_type, 'ТО3')
        self.assertEqual(self.parser_1.raw_data[1].place,
                         'Местоположение: Здание управления комплекса защитных сооружений')

        self.assertEqual(self.parser_2.raw_data[0].day, 17)
        self.assertEqual(self.parser_2.raw_data[0].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[0].place, 'С1 ПС 110/10кВ')
        self.assertEqual(self.parser_2.raw_data[2].day, 24)
        self.assertEqual(self.parser_2.raw_data[2].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[2].place, 'ПС 86')


class TestParserAskue(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb_askue = openpyxl.load_workbook(r'.\input data\Test Schedule Askue.xlsx')
        cls.sheet_1 = cls.wb_askue['февраль 8.1.36 ТО']
        cls.sheet_2 = cls.wb_askue['Февраль 10.4.36 ТО']
        cls.parser_1 = works_parser.ParserAskue(cls.sheet_1)
        cls.parser_2 = works_parser.ParserAskue(cls.sheet_2)

    # for j in parser_2.raw_data:
    #     print(j)

    def test_find_data_boundaries(self):
        self.assertEqual(self.parser_1._data_first_col, 6)
        self.assertEqual(self.parser_1._data_last_col, 33)
        self.assertEqual(self.parser_1._data_rows, [21, 22, 23])
        self.assertEqual(self.parser_1._days_row, 17)
        self.assertEqual(self.parser_2._data_first_col, 6)
        self.assertEqual(self.parser_2._data_last_col, 33)
        self.assertEqual(self.parser_2._data_rows, [22, 23, 24, 25, 26, 28, 29, 30, 31, 32, 34, 35, 36, 37, 38, 39])
        self.assertEqual(self.parser_2._days_row, 17)

    def test_find_month_year(self):
        self.assertEqual(self.parser_1.month_year, 'Февраль 2019 года')
        self.assertEqual(self.parser_2.month_year, 'Февраль 2019 года')

    def test_extract_jobs(self):
        last_1 = len(self.parser_1.raw_data) - 1
        self.assertEqual(len(self.parser_1.raw_data), 22)
        self.assertEqual(self.parser_1.raw_data[0].day, 8)
        self.assertEqual(self.parser_1.raw_data[0].work_type, 'ТО2')
        self.assertEqual(self.parser_1.raw_data[0].place,
                         'Местоподожение: Здание управления комплекса защитных сооружений')
        # "Местопо_д_ожение" не является опечаткой в тесте. Данная опечатка имеет место во входных данных
        self.assertEqual(self.parser_1.raw_data[last_1].day, 14)
        self.assertEqual(self.parser_1.raw_data[last_1].work_type, 'ТО2')
        self.assertEqual(self.parser_1.raw_data[last_1].place,
                         'Местоподожение: Здание управления комплекса защитных сооружений')

        last_2 = len(self.parser_2.raw_data) - 1
        self.assertEqual(len(self.parser_2.raw_data), 3)
        self.assertEqual(self.parser_2.raw_data[0].day, 13)
        self.assertEqual(self.parser_2.raw_data[0].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[0].place, 'ПС 223')
        self.assertEqual(self.parser_2.raw_data[1].day, 20)
        self.assertEqual(self.parser_2.raw_data[1].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[1].place, 'Котлин')
        self.assertEqual(self.parser_2.raw_data[2].day, 27)
        self.assertEqual(self.parser_2.raw_data[2].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[2].place, 'С1')


class TestParserTechReg(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb_tech_reg = openpyxl.load_workbook(r'.\input data\Test Schedule TechReg.xlsx')
        cls.sheet_1 = cls.wb_tech_reg['10.2.37 ТО']
        cls.sheet_2 = cls.wb_tech_reg['10.3.37 ТО']
        cls.parser_1 = works_parser.ParserTechReg(cls.sheet_1)
        cls.parser_2 = works_parser.ParserTechReg(cls.sheet_2)

    # for j in parser_2.raw_data:
    #     print(j)

    def test_find_data_boundaries(self):
        self.assertEqual(self.parser_1._data_first_col, 9)
        self.assertEqual(self.parser_1._data_last_col, 36)
        self.assertEqual(self.parser_1._data_rows, [x for x in range(10, 32)])
        self.assertEqual(self.parser_1._days_row, 7)
        self.assertEqual(self.parser_2._data_first_col, 9)
        self.assertEqual(self.parser_2._data_last_col, 39)
        self.assertEqual(self.parser_2._data_rows, [x for x in range(10, 44)])
        self.assertEqual(self.parser_2._days_row, 7)

    def test_find_month_year(self):
        self.assertEqual(self.parser_1.month_year, 'Февраль 2019 год')
        self.assertEqual(self.parser_2.month_year, 'Июль 2019 год')

    def test_extract_jobs(self):
        last_1 = len(self.parser_1.raw_data) - 1
        self.assertEqual(len(self.parser_1.raw_data), 2)
        self.assertEqual(self.parser_1.raw_data[0].day, 7)
        self.assertEqual(self.parser_1.raw_data[0].work_type, 'ТО2')
        self.assertEqual(self.parser_1.raw_data[0].place,
                         'Местоположение: Здание общеподстанционного управления 110 кВ ПС №360')
        self.assertEqual(self.parser_1.raw_data[last_1].day, 6)
        self.assertEqual(self.parser_1.raw_data[last_1].work_type, 'ТО2')
        self.assertEqual(self.parser_1.raw_data[last_1].place,
                         'Местоположение: Здание общеподстанционного управления 110 кВ ПС №360')

        self.assertEqual(len(self.parser_2.raw_data), 3)
        self.assertEqual(self.parser_2.raw_data[0].day, 9)
        self.assertEqual(self.parser_2.raw_data[0].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[0].place, 'Местоположение: Трансформаторная подстанция ПС С2 110/10 кВ')
        self.assertEqual(self.parser_2.raw_data[1].day, 10)
        self.assertEqual(self.parser_2.raw_data[1].work_type, 'ЕТО')
        self.assertEqual(self.parser_2.raw_data[1].place, 'Местоположение: Трансформаторная подстанция ПС С2 110/10 кВ')
        self.assertEqual(self.parser_2.raw_data[2].day, 10)
        self.assertEqual(self.parser_2.raw_data[2].work_type, 'ТО2')
        self.assertEqual(self.parser_2.raw_data[2].place, 'Местоположение: Трансформаторная подстанция ПС С2 110/10 кВ')


class TestPreProcessing(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        #     test_raw_places = open(r'.\input data\test raw places.txt')
        #     for line in test_raw_places:
        #         print(f'{line}  -->>  { extract_place_and_object(line)}')
        cls.wb_asu = openpyxl.load_workbook(r'.\input data\5. Графики на 05.18 АСУ.xlsx')
        cls.wb_test = openpyxl.load_workbook(r'.\input data\Test Schedule.xlsx')
        cls.sheet_asu_109 = cls.wb_asu['109. МОСТ']
        cls.sheet_asu_107 = cls.wb_asu['107. АСУ ТП']
        cls.sheet_test_107 = cls.wb_test['107. АСУ ТП']
        cls.parser_asu_109 = works_parser.ParserAsu(cls.sheet_asu_109)
        cls.parser_asu_107 = works_parser.ParserAsu(cls.sheet_asu_107)
        cls.parser_test_107 = works_parser.ParserAsu(cls.sheet_test_107)

    def test_find_num_in_str(self):
        func = pre_processing.find_num_in_str
        self.assertEqual(func('1'), 1)
        self.assertEqual(func('  85  \n'), 85)
        self.assertEqual(func('str str 43 str'), 43)
        self.assertEqual(func('72, 234'), 72)
        self.assertEqual(func('float 1,32'), 1)
        self.assertEqual(func(''), None)
        self.assertEqual(func('   '), None)
        self.assertEqual(func('str str'), None)
        # self.assertEqual(func(''), )
        # self.assertEqual(func(''), )

    def test_extract_system(self):
        self.assertEqual(pre_processing.find_system_by_sheet('107. АСУ ТП'), Systems.ASU_TP)
        self.assertEqual(pre_processing.find_system_by_sheet('108. АСУ И '), Systems.ASU_I)
        self.assertEqual(pre_processing.find_system_by_sheet('109. МОСТ'), Systems.ASU_AM)
        self.assertEqual(pre_processing.find_system_by_sheet('Лист 1'), None)
        self.assertEqual(pre_processing.find_system_by_sheet(''), None)

    def test_parser_to_jobs(self):
        jobs: List[pre_processing.Job] = []
        jobs.extend(pre_processing.parser_to_jobs(self.parser_test_107))
        last = len(jobs) - 1
        self.assertEqual(jobs[0].object, Objects.S1)
        self.assertEqual(jobs[0].place, 'С1 Север')
        self.assertEqual(jobs[0].work_type, 'ТО1')
        self.assertEqual(jobs[0].date, datetime.date(2018, 5, 1))
        self.assertEqual(jobs[0].system, Systems.ASU_TP)
        self.assertEqual(jobs[last].object, Objects.ZU)
        self.assertEqual(jobs[last].place, 'Здание управления КЗС')
        self.assertEqual(jobs[last].work_type, 'ТО4')
        self.assertEqual(jobs[last].date, datetime.date(2018, 5, 31))
        self.assertEqual(jobs[last].system, Systems.ASU_TP)

    def test_filter_work_type(self):
        self.assertEqual(pre_processing.filter_work_type('ТО1'), 'ТО1')  # Rus to Rus
        self.assertEqual(pre_processing.filter_work_type('TO2'), 'ТО2')  # Eng to Rus
        self.assertEqual(pre_processing.filter_work_type('ETO'), 'ЕТО')  # Eng to Rus
        self.assertEqual(pre_processing.filter_work_type('ЕТО \nТО1'), 'ЕТО \nТО1')  # Eng+Rus to Rus

    def test_extract_place_and_object(self):
        func = pre_processing.extract_place_and_object
        self.assertEqual(func('Местоположение: Здание общеподстанционного управления 110 кВ ПС №360'),
                         ('ПС 360', Objects.PS360))
        self.assertEqual(func('Судопропускное сооружение С1 Север ТП2'),
                         ('С1 Север ТП2', Objects.S1))
        self.assertEqual(func('Судопропускное сооружение С1 Север ПС 110/10 кВ'),
                         ('С1 ПС 110/10кВ', Objects.S1))
        self.assertEqual(func('Здание управления'),
                         ('Здание управления КЗС', Objects.ZU))
        self.assertEqual(func('ПТК ЗУ КЗС'),
                         ('Здание управления КЗС', Objects.ZU))
        self.assertEqual(func('Оборудование АСУ АМ'),
                         ('С2 АМ', Objects.S2))
        self.assertEqual(func('Водопропускное сооружение В-6'),
                         ('В6', Objects.V6))
        self.assertEqual(func('ПТК судопропускного сооружения - ПТК C1 север'),  # eng 'C'
                         ('С1 Север', Objects.S1))
        self.assertEqual(func('ПТК водопропускного сооружения ВЗ - ПТК ВЗ.'),
                         ('В3', Objects.V3))
        self.assertEqual(func('2. Котлин'),
                         ('Котлин', Objects.KOTLIN))
        self.assertEqual(func('1. Бронка'),
                         ('ПС 223', Objects.PS223))
        self.assertEqual(func('Местоподожение:  Здание трансформаторной подстанции '
                              '110/10 кВ ПС С1 судопропускного сооружения С-1'),
                         ('С1 ПС 110/10кВ', Objects.S1))
        self.assertEqual(func('Местоподожение: Трансформаторная подстанция ПС С2 110/10 кВ'),
                         ('С2 ПС 110/10кВ', Objects.S2))
        self.assertEqual(func('Местоподожение:Здание общеподстанционного управления 110 кВ  ПС №360'),
                         ('ПС 360', Objects.PS360))
        # self.assertEqual(func(''),
        #                  ('', ''))


class TestTableGenerator(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.style1 = NamedStyle(name='style1')

    def test_apply_style(self):
        test_out_wb = openpyxl.Workbook()
        ws = test_out_wb.active
        area = cell_styler.TableArea(first_row=3, last_row=6, first_col=2, last_col=4)
        cell_styler.apply_named_style(ws, self.style1, table_area=area)
        for row in range(area.first_row, area.last_row + 1):
            for col in range(area.first_col, area.last_col + 1):
                self.assertEqual(ws.cell(row, col).style, self.style1.name)


class TestApplicationFunctions(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb_test_asu = openpyxl.load_workbook(r'.\input data\Test Schedule.xlsx')
        cls.wb_test_vols = openpyxl.load_workbook(r'.\input data\Test Schedule VOLS.xlsx')

    def test_find_sheets_asu(self):
        sheets = application.find_sheets_asu(self.wb_test_asu)
        self.assertEqual(sheets[0].title, '107. АСУ ТП')
        self.assertEqual(sheets[1].title, '108. АСУ И ')
        self.assertEqual(sheets[2].title, '109. МОСТ')
        self.assertEqual(sheets[3].title, '110. ЛВС ')
        self.assertEqual(len(sheets), 4)

    def test_find_sheets_vols(self):
        sheets = application.find_sheets_vols(self.wb_test_vols)
        self.assertEqual(sheets[0].title, '8.1.38 ТО')
        self.assertEqual(sheets[1].title, '10.2.38 ТО')
        self.assertEqual(sheets[2].title, '10.3.38 ТО')
        self.assertEqual(sheets[3].title, '10.4.38 ТО')
        self.assertEqual(len(sheets), 4)


class TestDutySchedule(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.workbook = openpyxl.load_workbook(r'.\input data\Test график дежурств 19.02.xlsx')
        cls.schedule = duty_schedule.DutySchedule(cls.workbook.worksheets[0], duty_schedule.all_workers)

    def test_date2col(self):
        self.assertEqual(self.schedule._date2col(datetime.date(2019, 2, 1)), 3)
        self.assertEqual(self.schedule._date2col(datetime.date(2019, 2, 15)), 17)
        self.assertEqual(self.schedule._date2col(datetime.date(2019, 2, 28)), 30)

    def test_find_workers_row(self):
        workers = duty_schedule.team_s1.workers
        # print(duty_schedule.team_s1)
        correct_rows = [8, 9, 10, 11]
        self.assertEqual(len(workers), len(correct_rows))
        for i in range(0, len(workers)):
            # print(f'{self.schedule._find_worker_row(workers[i])} - {correct_rows[i]}')
            self.assertEqual(self.schedule._find_worker_row(workers[i]), correct_rows[i])

        workers = duty_schedule.team_s2.workers
        correct_rows = [14, 13]
        self.assertEqual(len(workers), len(correct_rows))
        for i in range(0, len(workers)):
            self.assertEqual(self.schedule._find_worker_row(workers[i]), correct_rows[i])

        workers = duty_schedule.team_v.workers
        correct_rows = [17, 18, 19]
        self.assertEqual(len(workers), len(correct_rows))
        for i in range(0, len(workers)):
            self.assertEqual(self.schedule._find_worker_row(workers[i]), correct_rows[i])

        workers = duty_schedule.team_vols.workers
        correct_rows = [23, 21, 22]
        self.assertEqual(len(workers), len(correct_rows))
        for i in range(0, len(workers)):
            self.assertEqual(self.schedule._find_worker_row(workers[i]), correct_rows[i])

        workers = duty_schedule.team_tk.workers
        correct_rows = [22, 21, 23]
        self.assertEqual(len(workers), len(correct_rows))
        for i in range(0, len(workers)):
            self.assertEqual(self.schedule._find_worker_row(workers[i]), correct_rows[i])

    def test_get_worker_status(self):
        w1 = duty_schedule.Worker('Гусев', 'Михаил', 'Владимирович', '+79675904368')
        w2 = duty_schedule.Worker('Мулин', 'Николай', 'Николаевич', '+79112283889')
        self.assertEqual(self.schedule.get_worker_status(w1, datetime.date(2019, 2, 1)),
                         duty_schedule.WorkerStatus.ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w1, datetime.date(2019, 2, 2)),
                         duty_schedule.WorkerStatus.ON_DUTY)
        self.assertEqual(self.schedule.get_worker_status(w1, datetime.date(2019, 2, 3)),
                         duty_schedule.WorkerStatus.NOT_ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w1, datetime.date(2019, 2, 4)),
                         duty_schedule.WorkerStatus.NOT_ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w1, datetime.date(2019, 2, 6)),
                         duty_schedule.WorkerStatus.NOT_ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w1, datetime.date(2019, 2, 7)),
                         duty_schedule.WorkerStatus.ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w2, datetime.date(2019, 2, 17)),
                         duty_schedule.WorkerStatus.NOT_ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w2, datetime.date(2019, 2, 18)),
                         duty_schedule.WorkerStatus.ON_DUTY)
        self.assertEqual(self.schedule.get_worker_status(w2, datetime.date(2019, 2, 19)),
                         duty_schedule.WorkerStatus.NOT_ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w2, datetime.date(2019, 2, 20)),
                         duty_schedule.WorkerStatus.NOT_ON_WORK)
        self.assertEqual(self.schedule.get_worker_status(w2, datetime.date(2019, 2, 22)),
                         duty_schedule.WorkerStatus.ON_WORK)

    def test_is_workday(self):
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 1)))
        self.assertFalse(self.schedule.is_workday(datetime.date(2019, 2, 2)))
        self.assertFalse(self.schedule.is_workday(datetime.date(2019, 2, 3)))
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 4)))
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 15)))
        self.assertFalse(self.schedule.is_workday(datetime.date(2019, 2, 16)))
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 21)))
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 22)))
        self.assertFalse(self.schedule.is_workday(datetime.date(2019, 2, 23)))
        self.assertFalse(self.schedule.is_workday(datetime.date(2019, 2, 24)))
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 25)))
        self.assertTrue(self.schedule.is_workday(datetime.date(2019, 2, 26)))

    def test_get_duty_str(self):
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 1)), 'Макаров В.В.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 2)), 'Гусев М.В.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 3)), 'Кушмылев Е.П.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 4)), 'Ильин А.В.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 11)), 'Подольский А.В.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 12)), 'Кушмылев Е.П.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 22)), 'Огородников А.Ю.')
        self.assertEqual(self.schedule.get_duty_str(datetime.date(2019, 2, 25)), 'Ильин А.В.')

    def test_get_performer(self):

        self.assertEqual(self.schedule.get_performer(duty_schedule.team_s1,
                                                     datetime.date(2019, 2, 1)).last_name, 'Харитонов')
        self.assertEqual(self.schedule.get_performer(duty_schedule.team_s1,
                                                     datetime.date(2019, 2, 2)).last_name, 'Гусев')
        self.assertEqual(self.schedule.get_performer(duty_schedule.team_s2,
                                                     datetime.date(2019, 2, 2)).last_name, 'Гусев')
        self.assertEqual(self.schedule.get_performer(duty_schedule.team_s1,
                                                     datetime.date(2019, 2, 8)).last_name, 'Гусев')


if __name__ == '__main__':
    unittest.main()
